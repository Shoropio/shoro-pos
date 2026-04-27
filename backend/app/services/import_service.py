from io import BytesIO
from decimal import Decimal

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.product import Product
from app.services.barcode_service import barcode_service


HEADER_ALIASES = {
    "internal_code": {"internal_code", "codigo", "codigo interno", "code"},
    "barcode": {"barcode", "codigo de barras", "código de barras"},
    "sku": {"sku"},
    "name": {"name", "nombre", "producto"},
    "cabys_code": {"cabys", "codigo cabys", "código cabys"},
    "purchase_price": {"purchase_price", "precio compra", "costo"},
    "sale_price": {"sale_price", "precio venta", "precio"},
    "tax_rate": {"tax_rate", "iva", "impuesto"},
    "stock": {"stock", "existencia", "inventario"},
    "min_stock": {"min_stock", "stock minimo", "stock mínimo"},
}


def _normalize(value: object) -> str:
    return str(value or "").strip().lower()


def _decimal(value: object, default: str = "0") -> Decimal:
    if value in (None, ""):
        return Decimal(default)
    return Decimal(str(value))


def import_products_excel(db: Session, content: bytes) -> dict:
    workbook = load_workbook(BytesIO(content), data_only=True)
    sheet = workbook.active
    raw_headers = [_normalize(cell.value) for cell in sheet[1]]
    mapping: dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        for index, header in enumerate(raw_headers):
            if header in aliases:
                mapping[field] = index
                break
    if "name" not in mapping or "sale_price" not in mapping:
        raise ValueError("El Excel requiere columnas nombre/name y precio/sale_price")

    created = updated = skipped = 0
    errors: list[str] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        try:
            def get(field: str, default=None):
                index = mapping.get(field)
                return row[index] if index is not None and index < len(row) else default

            internal_code = str(get("internal_code") or f"IMP-{row_number:06d}").strip()
            product = db.query(Product).filter(Product.internal_code == internal_code).first()
            values = {
                "internal_code": internal_code,
                "barcode": str(get("barcode") or barcode_service.generate_internal_barcode(internal_code)).strip(),
                "sku": str(get("sku") or internal_code).strip(),
                "name": str(get("name")).strip(),
                "cabys_code": str(get("cabys_code") or "").strip() or None,
                "purchase_price": _decimal(get("purchase_price")),
                "sale_price": _decimal(get("sale_price")),
                "tax_rate": _decimal(get("tax_rate"), "13"),
                "stock": _decimal(get("stock")),
                "min_stock": _decimal(get("min_stock")),
            }
            if product:
                for key, value in values.items():
                    setattr(product, key, value)
                updated += 1
            else:
                db.add(Product(**values))
                created += 1
        except Exception as exc:
            skipped += 1
            errors.append(f"Fila {row_number}: {exc}")
    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped, "errors": errors[:20]}
